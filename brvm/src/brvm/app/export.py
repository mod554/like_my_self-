"""Export tableur, et restitution texte qui n'exige aucune dépendance.

Deux sorties, une seule source : l'état assemblé. Elles montrent donc toujours
la même chose, et toutes deux commencent par le **bandeau de fraîcheur** — l'âge
de la donnée la plus ancienne employée. Un tableau de chiffres sans cette ligne
laisse croire qu'il décrit l'instant présent.

Le classeur est produit par ``openpyxl``, qui reste facultatif : sans lui, la
restitution texte donne la même information, et le message dit quoi installer.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Final

from brvm.app.api import MENTION_MARCHE
from brvm.app.etat import MOTIF_PERFORMANCE_ABSENTE, EtatSysteme
from brvm.domain.enums import MethodeValorisation
from brvm.domain.monnaie import format_xof
from brvm.market.allocation import Proposition
from brvm.market.criblage import Criblage
from brvm.utils.erreurs import ErreurConfiguration
from brvm.utils.journalisation import obtenir_journal

_journal = obtenir_journal("app.export")

#: Format d'affichage des montants en XOF. Pas de décimale : le franc ne circule
#: pas en centimes, et en afficher une laisserait croire à une précision absente.
FORMAT_XOF: Final[str] = '# ##0 "FCFA"'
FORMAT_POURCENT: Final[str] = "0.00%"

#: Mention portée par chaque feuille. Le système constate, il ne conseille pas.
MENTION: Final[str] = (
    "Constats produits par un outil de suivi. Aucun conseil d'investissement, "
    "aucune promesse de rendement. Vérifiez l'horodatage ci-dessus avant d'agir."
)


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dépend de l'installation
        raise ErreurConfiguration(
            "openpyxl n'est pas installé : l'export tableur est indisponible. "
            "Installez l'extra dédié (`pip install -e '.[tableur]'`), ou utilisez "
            "`restituer()` qui produit la même information en texte.",
        ) from exc
    return openpyxl


def _pourcent(valeur: Decimal | None) -> float | None:
    return float(valeur) if valeur is not None else None


def _ligne_performance(etat: EtatSysteme) -> str:
    """La performance, ou la raison précise de son absence.

    Elle n'est pas approchée : un rendement qu'on ne peut pas fonder ne se
    remplace pas par un nombre plausible.
    """
    resultat = etat.performance
    if resultat is None or resultat.valeur is None:
        return f"Performance : {etat.motif_performance_absente}"
    return (
        f"Performance (TWR, apports et retraits neutralisés) : {resultat.valeur:.2%} "
        f"sur {len(resultat.sous_periodes)} sous-période(s)"
    )


def restituer(etat: EtatSysteme) -> str:
    """Restitution texte complète, sans aucune dépendance.

    C'est la sortie de secours et la sortie de référence : tout ce que montrent
    le classeur et le tableau de bord se retrouve ici.
    """
    lignes: list[str] = [
        "=" * 78,
        f"ÉTAT DU PORTEFEUILLE — {etat.instant.isoformat()}",
        etat.entete_fraicheur(),
        "=" * 78,
        "",
        # `resume()` réaffiche le bandeau : il est déjà en tête, on ne le répète pas.
        "\n".join(etat.portefeuille.resume(etat.instant).splitlines()[1:]),
        "",
        _ligne_performance(etat),
        "",
        "-" * 78,
        "LIGNES",
        "-" * 78,
    ]
    for ligne in etat.portefeuille.lignes:
        if ligne.valorisee:
            lignes.append(
                f"{ligne.ticker:<8} {ligne.quantite:>8} × {format_xof(ligne.cours or 0):>14}"
                f" = {format_xof(ligne.valeur or 0):>16}"
                f"  (PRU {ligne.prix_revient_unitaire:.2f})"
            )
        else:
            lignes.append(
                f"{ligne.ticker:<8} {ligne.quantite:>8} titres — NON VALORISÉE : "
                f"{ligne.motif_indisponible}"
            )

    if etat.signaux:
        lignes += ["", "-" * 78, "SIGNAUX CONSTATÉS", "-" * 78]
        lignes += [f"  {signal.resume()}" for signal in etat.signaux[:20]]

    lignes += ["", "-" * 78, "RISQUE", "-" * 78, etat.risque.resume()]

    if etat.anomalies:
        lignes += ["", "-" * 78, "ANOMALIES OUVERTES", "-" * 78]
        lignes += [
            f"  [{anomalie.gravite.value}] {anomalie.source} — {anomalie.message}"
            for anomalie in etat.anomalies[:20]
        ]

    if etat.avertissements:
        lignes += ["", "-" * 78, "AVERTISSEMENTS", "-" * 78]
        lignes += [f"  • {message}" for message in etat.avertissements]

    lignes += ["", MENTION]
    return "\n".join(lignes)


def restituer_criblage(
    criblage: Criblage,
    propositions: Mapping[str, Proposition] | None = None,
    taille: int = 20,
) -> str:
    """Restitution texte du criblage de la cote.

    Les valeurs écartées figurent au même titre que les classées, avec leur
    raison. Sur cette place elles sont souvent la majorité : les omettre
    donnerait de la cote une image fausse, et ferait croire à une couverture
    que le système n'a pas.
    """
    lignes: list[str] = [
        "=" * 78,
        f"CRIBLAGE DE LA COTE — {criblage.instant.isoformat()}",
        criblage.entete_fraicheur(),
        f"Univers : {criblage.univers} valeur(s), {len(criblage.analyses)} analysée(s), "
        f"{criblage.fondamentaux_renseignes} avec fondamentaux saisis",
        "=" * 78,
    ]

    for classement in criblage.classements.values():
        lignes += [
            "",
            "-" * 78,
            f"{classement.libelle.upper()} — {classement.couverture_cote} classée(s)",
            "-" * 78,
            f"  {classement.description.strip()}",
            "",
        ]
        tete = classement.tete(taille)
        for place, rang in enumerate(tete, start=1):
            cours = format_xof(rang.analyse.cours) if rang.analyse.cours else "cours absent"
            lignes.append(
                f"  {place:>2}. {rang.valeur:.3f}  {rang.ticker:<8} {cours:>16}"
                f"   couverture {rang.score.couverture:.0%}"
            )
        reste = len(classement.classes) - len(tete)
        if reste > 0:
            # Jamais de troncature muette : le nombre omis et le réglage qui l'a
            # décidé figurent tous les deux.
            lignes.append(
                f"   … {reste} valeur(s) classée(s) de plus, non affichée(s) — "
                f"`analyse.taille_classement` vaut {taille}."
            )
        for rang in classement.ecartes:
            lignes.append(f"   ——      {rang.ticker:<8} {rang.score.motif_absent}")
        for message in classement.avertissements:
            lignes.append(f"  ⚠ {message}")

    for nom, proposition in (propositions or {}).items():
        libelle = criblage.classements[nom].libelle if nom in criblage.classements else nom
        lignes += [
            "",
            "-" * 78,
            f"RÉPARTITION POSSIBLE — {libelle.upper()}",
            "-" * 78,
            proposition.resume(),
        ]
        for ecarte in proposition.ecartes:
            lignes.append(f"   ——      {ecarte.ticker:<8} {ecarte.motif}")
        for message in proposition.avertissements:
            lignes.append(f"  ⚠ {message}")

    if criblage.ecartees:
        lignes += ["", "-" * 78, "VALEURS NON ANALYSABLES", "-" * 78]
        lignes += [f"  {e.ticker:<8} {e.motif}" for e in criblage.ecartees]

    if criblage.avertissements:
        lignes += ["", "-" * 78, "AVERTISSEMENTS", "-" * 78]
        lignes += [f"  • {message}" for message in criblage.avertissements]

    lignes += ["", MENTION_MARCHE, MENTION]
    return "\n".join(lignes)


class _Classeur:
    """Petit enrobage d'openpyxl : chaque feuille naît avec son bandeau."""

    def __init__(self, etat: EtatSysteme) -> None:
        self.etat = etat
        self.module = _openpyxl()
        self.classeur = self.module.Workbook()
        self.classeur.remove(self.classeur.active)

    def feuille(self, titre: str, entetes: Sequence[str]) -> Any:
        feuille = self.classeur.create_sheet(titre[:31])
        feuille.append([f"État au {self.etat.instant.isoformat()}"])
        feuille.append([self.etat.entete_fraicheur()])
        feuille.append([MENTION])
        feuille.append([])
        feuille.append(list(entetes))
        for cellule in feuille[5]:
            cellule.font = self.module.styles.Font(bold=True)
        feuille.freeze_panes = "A6"
        return feuille

    def ajuster(self, feuille: Any) -> None:
        for colonne in feuille.columns:
            largeur = max((len(str(cellule.value or "")) for cellule in colonne), default=10)
            feuille.column_dimensions[colonne[0].column_letter].width = min(60, largeur + 2)


def exporter(etat: EtatSysteme, chemin: Path | str) -> Path:
    """Écrit le classeur et renvoie son chemin.

    Cinq feuilles : positions, signaux, risque, anomalies, collectes. Chacune
    répond à une question, et chacune porte l'horodatage de la donnée la plus
    ancienne employée.
    """
    cible = Path(chemin).expanduser()
    classeur = _Classeur(etat)

    _feuille_positions(classeur)
    _feuille_signaux(classeur)
    _feuille_risque(classeur)
    _feuille_anomalies(classeur)
    _feuille_collectes(classeur)

    cible.parent.mkdir(parents=True, exist_ok=True)
    classeur.classeur.save(cible)
    _journal.info(
        "Export tableur écrit",
        extra={"chemin": str(cible), "lignes": len(etat.portefeuille.lignes)},
    )
    return cible


def _feuille_positions(classeur: _Classeur) -> None:
    etat = classeur.etat
    feuille = classeur.feuille(
        "Positions",
        [
            "Valeur",
            "Quantité",
            "Prix de revient unitaire",
            "Coût total",
            "Cours",
            "Séance du cours",
            "Âge du cours (min)",
            "Valeur",
            "Poids",
            "Plus-value latente brute",
            "Plus-value latente nette",
            "Frais de cession estimés",
            "Impôt estimé",
            "Dividendes nets",
            "Motif si non valorisée",
        ],
    )
    for ligne in etat.portefeuille.lignes:
        age = ligne.age_minutes(etat.instant)
        feuille.append(
            [
                ligne.ticker,
                ligne.quantite,
                float(ligne.prix_revient_unitaire),
                ligne.cout_total,
                ligne.cours,
                ligne.date_cours.isoformat() if ligne.date_cours else None,
                float(age) if age is not None else None,
                ligne.valeur,
                _pourcent(ligne.poids),
                ligne.plus_value_latente_brute,
                ligne.plus_value_latente_nette,
                ligne.frais_cession_estimes,
                ligne.impot_estime,
                ligne.dividendes_nets,
                ligne.motif_indisponible,
            ]
        )
    for rangee in feuille.iter_rows(min_row=6):
        for position in (4, 5, 8, 10, 11, 12, 13, 14):
            rangee[position - 1].number_format = FORMAT_XOF
        rangee[8].number_format = FORMAT_POURCENT

    feuille.append([])
    feuille.append(["TOTAL coût engagé", etat.portefeuille.cout_total])
    feuille.append(["TOTAL valeur", etat.portefeuille.valeur_totale])
    feuille.append(["Plus-value latente brute", etat.portefeuille.plus_value_latente_brute])
    feuille.append(["Plus-value latente nette", etat.portefeuille.plus_value_latente_nette])
    feuille.append(["Dividendes nets encaissés", etat.portefeuille.dividendes_nets_encaisses])
    feuille.append(["Performance", MOTIF_PERFORMANCE_ABSENTE])
    # PMP et FIFO répondent à deux questions : les deux figurent, jamais un seul
    # chiffre présenté comme « la » plus-value.
    for methode in (MethodeValorisation.PMP, MethodeValorisation.FIFO):
        suivi = etat.suivis.get(methode)
        if suivi is None:
            continue
        realisees = sum(cession.plus_value_brute for cession in suivi.cessions)
        feuille.append([f"Plus-values réalisées ({methode.value})", realisees])
    classeur.ajuster(feuille)


def _feuille_signaux(classeur: _Classeur) -> None:
    feuille = classeur.feuille(
        "Signaux",
        [
            "Valeur",
            "Sens",
            "Règle",
            "Séance de constat",
            "Exécutable à partir du",
            "Confiance",
            "Niveau",
            "Explication",
            "Avertissements",
        ],
    )
    for signal in classeur.etat.signaux:
        feuille.append(
            [
                signal.ticker,
                signal.sens.value,
                signal.regle,
                signal.date_constat.isoformat(),
                # La date d'exécution est une colonne à part entière : un signal
                # constaté à la clôture n'était pas exécutable pendant la séance.
                signal.date_execution.isoformat(),
                float(signal.confiance.valeur),
                signal.confiance.niveau,
                signal.explication,
                " ".join(signal.avertissements),
            ]
        )
    classeur.ajuster(feuille)


def _feuille_risque(classeur: _Classeur) -> None:
    rapport = classeur.etat.risque
    feuille = classeur.feuille(
        "Risque", ["Nature", "Objet", "Constat", "Mesure", "Limite", "Respecté"]
    )
    for constat in rapport.concentrations:
        feuille.append(
            [
                "Concentration",
                f"{constat.dimension.value} {constat.cle}",
                constat.resume(),
                _pourcent(constat.poids),
                _pourcent(constat.limite),
                "oui" if constat.respecte else "NON",
            ]
        )
    for liquidite in rapport.liquidites:
        feuille.append(
            [
                "Liquidité",
                liquidite.ticker,
                liquidite.resume(),
                float(liquidite.seances_pour_deboucler)
                if liquidite.seances_pour_deboucler is not None
                else None,
                None,
                "non mesurable" if not liquidite.mesurable else "",
            ]
        )
    for stop in rapport.stops:
        feuille.append(["Stop ATR", stop.ticker, stop.resume(), stop.niveau, None, ""])
    for message in rapport.avertissements:
        feuille.append(["Avertissement", "", message, None, None, ""])
    for rangee in feuille.iter_rows(min_row=6):
        if rangee[0].value == "Concentration":
            rangee[3].number_format = FORMAT_POURCENT
            rangee[4].number_format = FORMAT_POURCENT
    classeur.ajuster(feuille)


def _feuille_anomalies(classeur: _Classeur) -> None:
    feuille = classeur.feuille(
        "Anomalies",
        ["Détectée le", "Gravité", "Source", "Type", "Valeur", "Séance", "Message"],
    )
    for anomalie in classeur.etat.anomalies:
        feuille.append(
            [
                anomalie.detectee_le.isoformat(),
                anomalie.gravite.value,
                anomalie.source,
                anomalie.type_anomalie,
                anomalie.ticker,
                anomalie.date_seance.isoformat() if anomalie.date_seance else None,
                anomalie.message,
            ]
        )
    classeur.ajuster(feuille)


def _feuille_collectes(classeur: _Classeur) -> None:
    feuille = classeur.feuille("Collectes", ["Journal des collectes"])
    for ligne in classeur.etat.journal_collectes:
        feuille.append([ligne])
    classeur.ajuster(feuille)


def horodater(chemin: Path | str, instant: datetime) -> Path:
    """Ajoute l'horodatage au nom du fichier : un export non daté est ininterprétable."""
    cible = Path(chemin)
    marque = instant.strftime("%Y%m%d-%H%M")
    return cible.with_name(f"{cible.stem}_{marque}{cible.suffix or '.xlsx'}")
