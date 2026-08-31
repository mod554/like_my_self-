"""Connecteur de secours : fichier CSV ou Excel alimenté à la main.

C'est le seul connecteur pleinement opérationnel dès l'installation, et c'est
volontaire : il ne dépend d'aucune source extérieure, d'aucune structure de page
et d'aucune condition d'utilisation. Il permet de faire tourner tout le système
— indicateurs, portefeuille, backtest — sur des cotations que vous saisissez ou
exportez vous-même, avant même qu'un connecteur réseau ait été vérifié.

Le format décrit ici est **celui du système**, pas celui d'un tiers : il peut
donc être spécifié sans rien supposer d'une source extérieure.

Colonnes attendues (l'ordre est libre, les en-têtes font foi) :

=========================  ============================================================
``ticker``                 obligatoire
``date_seance``            obligatoire, AAAA-MM-JJ
``statut_seance``          COTEE, SANS_TRANSACTION, SUSPENDUE, FERMEE ou INCONNU
``ouverture``              entier XOF, vide si non publié
``plus_haut``              idem
``plus_bas``               idem
``cloture``                idem
``cours_precedent``        idem
``volume_titres``          entier ; 0 seulement si vous savez qu'il n'y a pas eu d'échange
``volume_xof``             entier, vide si non publié
``nb_transactions``        entier, vide si non publié
``limite_achat``           meilleure limite à l'achat, vide si non publiée
``limite_vente``           meilleure limite à la vente, vide si non publiée
``horodatage_donnee``      ISO 8601 avec fuseau ; à défaut, la clôture de la séance
``commentaire``            libre
=========================  ============================================================

Séparateur ``,``, ``;`` ou tabulation détecté automatiquement. Les espaces de
milliers sont tolérés. Une cellule vide vaut « non publié », jamais zéro. Les
lignes commençant par ``#`` sont des commentaires.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any, Final

from brvm.config.modeles import ConfigSource, Configuration
from brvm.ingestion.base import DataSource, LigneCollectee, ResultatCollecte, statut_depuis_lignes
from brvm.ingestion.conversion import CHAMPS, CHAMPS_OBLIGATOIRES, ConvertisseurCotation, nettoyer
from brvm.utils.erreurs import ErreurSource
from brvm.utils.journalisation import obtenir_journal

_journal = obtenir_journal("ingestion.fichier")

EXTENSIONS_TABLEUR: Final[frozenset[str]] = frozenset({".xlsx", ".xlsm"})

#: En-tête du fichier modèle.
MODELE_ENTETE: Final[str] = ",".join([*CHAMPS, "horodatage_donnee"])


class SourceFichier(DataSource):
    """Lit les cotations depuis un fichier local, sans aucun accès réseau."""

    def __init__(self, source: ConfigSource, configuration: Configuration) -> None:
        if source.chemin_fichier is None:
            raise ErreurSource(
                f"La source {source.nom!r} est de type fichier mais aucun chemin_fichier "
                "n'est renseigné dans la configuration.",
                source=source.nom,
            )
        self.nom = source.nom
        self.source = source
        self.configuration = configuration
        self.chemin = Path(source.chemin_fichier)
        self._convertisseur = ConvertisseurCotation(self.nom, configuration)

    # ------------------------------------------------------------------ contrat

    def disponible(self) -> bool:
        return self.chemin.is_file()

    def collecter(self, jour: date | None = None) -> ResultatCollecte:
        debut = self.maintenant()
        if not self.disponible():
            return self.echec(
                debut,
                "Fichier de cotations introuvable. Créez-le avec "
                "`SourceFichier.ecrire_modele`, ou exportez-le depuis votre SGI.",
                origine=str(self.chemin),
            )
        try:
            lignes_brutes = list(self._lire())
        except ErreurSource as exc:
            return self.echec(debut, exc.message, origine=str(self.chemin))
        except (OSError, UnicodeDecodeError) as exc:
            return self.echec(debut, f"Fichier illisible : {exc}", origine=str(self.chemin))

        collectees: list[LigneCollectee] = []
        avertissements: list[str] = []
        collecte = self.maintenant()

        for numero, brut in lignes_brutes:
            if jour is not None and nettoyer(brut.get("date_seance")) != jour.isoformat():
                continue
            horodatage = self._horodatage_declare(brut, numero)
            if isinstance(horodatage, str):
                collectees.append(LigneCollectee(brut={"ligne": numero, **brut}, erreur=horodatage))
                continue
            ligne, avertissement = self._convertisseur.convertir(
                brut, collecte, horodatage_donnee=horodatage, repere=f"ligne {numero}"
            )
            collectees.append(ligne)
            if avertissement:
                avertissements.append(avertissement)

        return ResultatCollecte(
            source=self.nom,
            statut=statut_depuis_lignes(collectees),
            debut=debut,
            fin=self.maintenant(),
            lignes=tuple(collectees),
            origine=str(self.chemin),
            avertissements=tuple(avertissements),
            message=f"{len(collectees)} ligne(s) lue(s) depuis {self.chemin.name}",
        )

    # ------------------------------------------------------------------ lecture

    def _lire(self) -> Iterator[tuple[int, dict[str, Any]]]:
        if self.chemin.suffix.lower() in EXTENSIONS_TABLEUR:
            yield from self._lire_tableur()
        else:
            yield from self._lire_csv()

    def _lire_csv(self) -> Iterator[tuple[int, dict[str, Any]]]:
        contenu = self.chemin.read_text(encoding="utf-8-sig")
        utiles = [
            ligne
            for ligne in contenu.splitlines()
            if ligne.strip() and not ligne.lstrip().startswith("#")
        ]
        if not utiles:
            raise ErreurSource("Fichier vide ou entièrement commenté.", source=self.nom)
        try:
            dialecte: Any = csv.Sniffer().sniff(utiles[0], delimiters=",;\t")
        except csv.Error:
            dialecte = csv.get_dialect("excel")
        lecteur = csv.DictReader(utiles, dialect=dialecte)
        self._verifier_entetes(lecteur.fieldnames)
        for numero, brut in enumerate(lecteur, start=2):
            yield numero, {cle: valeur for cle, valeur in brut.items() if cle is not None}

    def _lire_tableur(self) -> Iterator[tuple[int, dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dépend de l'installation
            raise ErreurSource(
                "Lecture de classeur Excel demandée mais openpyxl n'est pas installé. "
                'Installez `pip install ".[tableur]"`, ou exportez le fichier en CSV.',
                source=self.nom,
            ) from exc
        classeur = load_workbook(self.chemin, read_only=True, data_only=True)
        try:
            feuille = classeur.active
            if feuille is None:
                raise ErreurSource("Classeur sans feuille active.", source=self.nom)
            iterateur = feuille.iter_rows(values_only=True)
            try:
                entetes = [nettoyer(cellule) for cellule in next(iterateur)]
            except StopIteration as exc:
                raise ErreurSource("Classeur vide.", source=self.nom) from exc
            self._verifier_entetes(entetes)
            for numero, valeurs in enumerate(iterateur, start=2):
                if all(cellule is None for cellule in valeurs):
                    continue
                yield numero, dict(zip(entetes, valeurs, strict=False))
        finally:
            classeur.close()

    def _verifier_entetes(self, entetes: Sequence[str | None] | None) -> None:
        presents = {nettoyer(entete) for entete in (entetes or []) if nettoyer(entete)}
        manquants = CHAMPS_OBLIGATOIRES - presents
        if manquants:
            raise ErreurSource(
                "Colonnes obligatoires absentes du fichier : "
                + ", ".join(sorted(manquants))
                + f". En-tête attendu : {MODELE_ENTETE}",
                source=self.nom,
                fichier=str(self.chemin),
            )
        inconnues = presents - set(CHAMPS) - {"horodatage_donnee"}
        if inconnues:
            _journal.info(
                "Colonnes ignorées dans le fichier de cotations",
                extra={"colonnes": sorted(inconnues), "fichier": str(self.chemin)},
            )

    @staticmethod
    def _horodatage_declare(brut: dict[str, Any], numero: int) -> datetime | str | None:
        """Lit ``horodatage_donnee`` s'il est présent. Renvoie le message d'erreur sinon."""
        declare = nettoyer(brut.get("horodatage_donnee"))
        if not declare:
            return None
        try:
            horodatage = datetime.fromisoformat(declare)
        except ValueError:
            return (
                f"ligne {numero} rejetée — horodatage_donnee « {declare} » n'est pas une "
                "date ISO 8601."
            )
        if horodatage.tzinfo is None:
            return (
                f"ligne {numero} rejetée — horodatage_donnee doit porter un fuseau "
                "(exemple : 2026-03-02T15:30:00+00:00). Sans fuseau, l'âge de la donnée "
                "est incalculable."
            )
        return horodatage

    # ---------------------------------------------------------------- utilitaire

    @staticmethod
    def ecrire_modele(chemin: Path) -> Path:
        """Écrit un fichier modèle commenté, prêt à être rempli."""
        chemin = Path(chemin)
        chemin.parent.mkdir(parents=True, exist_ok=True)
        chemin.write_text(
            "# Cotations saisies ou exportées manuellement.\n"
            "# Une cellule vide signifie « non publié », jamais zéro.\n"
            "# statut_seance : COTEE, SANS_TRANSACTION, SUSPENDUE, FERMEE ou INCONNU.\n"
            "# Laisser statut_seance vide avec un volume nul donne INCONNU : le système ne\n"
            "# suppose pas qu'il s'agit d'une séance sans transaction.\n"
            f"{MODELE_ENTETE}\n",
            encoding="utf-8",
        )
        return chemin
